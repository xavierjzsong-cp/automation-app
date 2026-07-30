"""End-to-end check for the replaceable coating mapper service boundary."""

from __future__ import annotations

from pathlib import Path
import sys
from tempfile import TemporaryDirectory
from typing import Any

import fitz
from openpyxl import Workbook, load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from src.mappers.coating_mapper import CoatingMapper  # noqa: E402
from src.services.template_generation_service import (  # noqa: E402
    GenerationRequest,
    TemplateGenerationService,
)


EXPECTED_COATING = {
    "top_thread_coating": "CSP-83",
    "bottom_thread_coating": "CSP-83",
    "body_coating": "CSP-99",
}


class RecordingCoatingMapper(CoatingMapper):
    """Real coating mapper that records calls at the replaceable boundary."""

    def __init__(self) -> None:
        self.calls: list[dict[str, Any]] = []

    def build_mapped_data(
        self,
        router_output: dict[str, Any],
    ) -> dict[str, str | None]:
        self.calls.append(router_output)
        return super().build_mapped_data(router_output)


def main() -> None:
    text = (
        "POTS Document number: 123 Rev: A\n"
        "CP Part Number ABC-001\n"
        "Product Description Pup Joint 13CR(80) 5.5 17# VAM TOP BOX X "
        "5.5 17# TSH WEDGE PIN OAL 120\n"
        "ANSI/NACE MR0175/ISO 15156 (Yes/No) Yes\n"
        "QCP (Standard/Client Specific) Standard\n"
    )

    with TemporaryDirectory() as tmp_name:
        root = Path(tmp_name)
        pdf = root / "input.pdf"
        template = root / "template.xlsx"
        output_dir = root / "out"

        document = fitz.open()
        page = document.new_page()
        page.insert_text((72, 72), text)
        document.save(pdf)
        document.close()

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Target"
        workbook.save(template)

        coating_mapper = RecordingCoatingMapper()
        service = TemplateGenerationService(coating_mapper=coating_mapper)
        request = GenerationRequest(
            input_path=pdf,
            template_path=template,
            output_dir=output_dir,
            target_sheet_name="Target",
            user_name="Tester",
        )

        result = service.generate(request)
        output = Path(result.output_file)
        assert output.exists()

        assert len(coating_mapper.calls) == 1
        assert coating_mapper.calls[0] == result.routing_result
        assert result.coating_data == EXPECTED_COATING
        assert result.writer_result["formatted"]["coating"] == EXPECTED_COATING

        workbook = load_workbook(output)
        sheet = workbook["Target"]
        assert sheet["B29"].value == "CSP-83"
        assert sheet["B31"].value == "CSP-83"
        assert sheet["B32"].value == "CSP-99"
        workbook.close()

    print("service coating flow ok")


if __name__ == "__main__":
    main()
