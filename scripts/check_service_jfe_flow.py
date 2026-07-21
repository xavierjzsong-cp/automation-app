"""Smoke and repeatability checks for JFE service integration."""

from __future__ import annotations

import sys
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any

import fitz
from openpyxl import Workbook, load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from src.services.template_generation_service import (  # noqa: E402
    GenerationRequest,
    TemplateGenerationService,
)


EXPECTED_RESULT = {
    "tensile": "561000",
    "compression": "540000",
    "burst": "12345",
    "collapse": "10987",
    "drift": "4.767",
    "od": {"nominal": "6.125", "tol_1": "+0.015", "tol_2": "-0.007"},
    "id": {"nominal": "4.601", "tol_1": "+0.010", "tol_2": "-0.005"},
    "external_length": "4.875",
    "internal_length": "4.875",
}


class FakeJfeAdapter:
    """Local adapter boundary used without partner-site network traffic."""

    instances: list["FakeJfeAdapter"] = []

    def __init__(
        self,
        base_url: str,
        datasheet_url: str,
        blanking_url: str,
        logs_dir: str | Path,
        headless: bool,
        slow_mo: int,
        timeout_ms: int,
        navigation_timeout_ms: int,
    ) -> None:
        self.base_url = base_url
        self.datasheet_url = datasheet_url
        self.blanking_url = blanking_url
        self.logs_dir = Path(logs_dir)
        self.headless = headless
        self.slow_mo = slow_mo
        self.timeout_ms = timeout_ms
        self.navigation_timeout_ms = navigation_timeout_ms
        self.closed = False
        self.run_calls: list[dict[str, Any]] = []
        self.instances.append(self)

    def run(self, mapped_result: dict[str, Any]) -> dict[str, Any]:
        self.run_calls.append(mapped_result)
        return {
            key: value.copy() if isinstance(value, dict) else value
            for key, value in EXPECTED_RESULT.items()
        }

    def close(self) -> None:
        self.closed = True


def check_repeated_dispatch(
    service: TemplateGenerationService,
    mapped_result: dict[str, Any],
) -> None:
    """Exercise the replaceable service-to-adapter boundary repeatedly."""
    partners_config = service._load_partners_config(service.partners_config_path)

    for _ in range(250):
        result = service._run_adapter_for_mapped_result(
            mapped_result=mapped_result,
            partners_config=partners_config,
            show_browser=False,
        )
        assert result == EXPECTED_RESULT


def main() -> None:
    FakeJfeAdapter.instances.clear()
    text = (
        "POTS Document number: 123 Rev: A\n"
        "CP Part Number ABC-001\n"
        "Product Description Pup Joint 13CR(80) 5.5 17# JFE BEAR BOX X "
        "5.5 17# VAM TOP PIN OAL 120\n"
        "ANSI/NACE MR0175/ISO 15156 (Yes/No) Yes\n"
        "QCP (Standard/Client Specific) Standard\n"
    )

    with TemporaryDirectory() as tmp_name:
        root = Path(tmp_name)
        pdf = root / "input.pdf"
        template = root / "template.xlsx"
        output_dir = root / "out"
        logs_dir = root / "logs"

        doc = fitz.open()
        page = doc.new_page()
        page.insert_text((72, 72), text)
        doc.save(pdf)
        doc.close()

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Target"
        workbook.save(template)

        service = TemplateGenerationService(
            adapter_factories={"JFE": FakeJfeAdapter},
            logs_dir=logs_dir,
        )
        request = GenerationRequest(
            input_path=pdf,
            template_path=template,
            output_dir=output_dir,
            target_sheet_name="Target",
            user_name="Tester",
            run_partner_adapters=True,
        )

        result = service.generate(request)
        output = Path(result.output_file)
        assert output.exists()

        assert len(FakeJfeAdapter.instances) == 1
        adapter = FakeJfeAdapter.instances[0]
        assert adapter.closed is True
        assert adapter.base_url == "https://www.jfetools.com/"
        assert adapter.datasheet_url == "https://www.jfetools.com/datasheet_generator"
        assert adapter.blanking_url == "https://www.jfetools.com/blanking_dimensions"
        assert adapter.logs_dir == logs_dir
        assert adapter.headless is True
        assert adapter.slow_mo == 300
        assert adapter.timeout_ms == 10000
        assert adapter.navigation_timeout_ms == 60000

        assert len(adapter.run_calls) == 1
        mapped = adapter.run_calls[0]
        assert mapped["partner"] == "JFE"
        assert mapped["side"] == "upper"
        assert mapped["drift_extraction"] is True
        assert mapped["connection"] == {
            "name": "JFEBEAR",
            "od": "5.500",
            "weight": "17",
            "material_family": "13CR",
            "yield_strength": "80",
            "grade_source": "standard",
            "friction": "API Modified",
            "coupling": "STD",
            "type": "BOX",
        }

        assert result.top_adapter == EXPECTED_RESULT
        assert result.bottom_adapter is None
        assert [item["partner"] for item in result.mapped_results or []] == ["JFE", "VAM"]

        formatted = result.writer_result["formatted"]
        assert formatted["top_thread"]["burst"] == "12345"
        assert "burst" not in formatted["bottom_thread"]

        workbook = load_workbook(output)
        sheet = workbook["Target"]
        assert sheet["B6"].value == "ABC-001"
        assert sheet["B28"].value == "5.5 - 17# JFE BEAR BOX"
        assert sheet["B30"].value == "5.5 - 17# VAM TOP PIN"
        assert sheet["B13"].value == "6.140"
        assert sheet["B14"].value == "4.596"
        assert sheet["B15"].value == "120.125"
        assert sheet["B22"].value == "561000K"
        assert sheet["B23"].value == "540000K"
        assert sheet["B24"].value == "12,345"
        assert sheet["B25"].value == "10,987"
        assert sheet["B33"].value == "NA"
        assert sheet["H13"].value == "6.125 +.015 /-.007"
        assert sheet["H14"].value == "4.601 +.010 /-.005"
        assert sheet["H15"].value == "4.875 +.125/ -.000"
        assert sheet["H17"].value == "4.875 +.125/ -.000"
        workbook.close()

        check_repeated_dispatch(service, mapped)

    repeated_instances = FakeJfeAdapter.instances[1:]
    assert len(repeated_instances) == 250
    assert all(instance.closed for instance in repeated_instances)
    assert all(instance.run_calls == [mapped] for instance in repeated_instances)

    print("service jfe flow ok")


if __name__ == "__main__":
    main()
