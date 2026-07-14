"""Smoke check for TSH adapter integration in the generation service."""

from __future__ import annotations

import sys
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any

import fitz
from openpyxl import Workbook, load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from src.services.template_generation_service import (  # noqa: E402
    GenerationRequest,
    TemplateGenerationService,
)


class FakeTshAdapter:
    instances: list["FakeTshAdapter"] = []

    def __init__(
        self,
        base_url: str,
        datasheet_url: str,
        blanking_url: str,
        logs_dir: str | Path,
        headless: bool,
        slow_mo: int,
        timeout_ms: int,
        navigation_timeout_ms: int,
    ) -> None:
        self.base_url = base_url
        self.datasheet_url = datasheet_url
        self.blanking_url = blanking_url
        self.logs_dir = Path(logs_dir)
        self.headless = headless
        self.slow_mo = slow_mo
        self.timeout_ms = timeout_ms
        self.navigation_timeout_ms = navigation_timeout_ms
        self.closed = False
        self.run_calls: list[dict[str, Any]] = []
        self.instances.append(self)

    def run(self, mapped_result: dict[str, Any]) -> dict[str, Any]:
        self.run_calls.append(mapped_result)
        return {
            "tensile": "242",
            "compression": "296",
            "burst": "7740",
            "collapse": "6290",
            "od": {"min": "5.490", "max": "5.510"},
            "id": {"min": "4.792", "max": "4.812"},
            "external_length": "3.400",
            "internal_length": "3.400",
            "drift": "4.767",
        }

    def close(self) -> None:
        self.closed = True


def main() -> None:
    FakeTshAdapter.instances.clear()
    text = (
        "POTS Document number: 123 Rev: A\n"
        "CP Part Number ABC-001\n"
        "Product Description Pup Joint 13CR(80) 5.5 17# VAM TOP BOX X "
        "5.5 17# TSH WEDGE PIN OAL 120\n"
        "ANSI/NACE MR0175/ISO 15156 (Yes/No) Yes\n"
        "QCP (Standard/Client Specific) Standard\n"
    )

    tmp = TemporaryDirectory()
    root = Path(tmp.name)
    pdf = root / "input.pdf"
    template = root / "template.xlsx"
    output_dir = root / "out"
    logs_dir = root / "logs"

    doc = fitz.open()
    page = doc.new_page()
    page.insert_text((72, 72), text)
    doc.save(pdf)
    doc.close()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Target"
    workbook.save(template)

    service = TemplateGenerationService(
        adapter_factories={"TSH": FakeTshAdapter},
        logs_dir=logs_dir,
    )
    request = GenerationRequest(
        input_path=pdf,
        template_path=template,
        output_dir=output_dir,
        target_sheet_name="Target",
        user_name="Tester",
        run_partner_adapters=True,
    )

    result = service.generate(request)
    output = Path(result.output_file)
    assert output.exists()

    assert len(FakeTshAdapter.instances) == 1
    adapter = FakeTshAdapter.instances[0]
    assert adapter.closed is True
    assert adapter.base_url == "https://dcp.tenaris.com/en"
    assert adapter.datasheet_url == "https://dcp.tenaris.com/Product_Datasheet"
    assert adapter.blanking_url == "https://dcp.tenaris.com/BlankingDimensions"
    assert adapter.headless is True
    assert adapter.logs_dir == logs_dir
    assert adapter.navigation_timeout_ms == 60000

    assert len(adapter.run_calls) == 1
    mapped = adapter.run_calls[0]
    assert mapped["partner"] == "TSH"
    assert mapped["side"] == "lower"
    assert mapped["connection"]["name"] == "WEDGE"
    assert mapped["connection"]["od"] == "5.500"
    assert mapped["connection"]["weight"] == "17.00"

    assert result.top_adapter is None
    assert result.bottom_adapter == {
        "tensile": "242",
        "compression": "296",
        "burst": "7740",
        "collapse": "6290",
        "od": {"min": "5.490", "max": "5.510"},
        "id": {"min": "4.792", "max": "4.812"},
        "external_length": "3.400",
        "internal_length": "3.400",
        "drift": "4.767",
    }
    assert [item["partner"] for item in result.mapped_results or []] == ["VAM", "TSH"]

    formatted = result.writer_result["formatted"]
    assert "burst" not in formatted["top_thread"]
    assert formatted["bottom_thread"]["burst"] == "7740"
    assert formatted["bottom_thread"]["drift"] == "4.767"

    workbook = load_workbook(output)
    sheet = workbook["Target"]
    assert sheet["B6"].value == "ABC-001"
    assert sheet["B28"].value == "5.5 - 17# VAM TOP BOX"
    assert sheet["B30"].value == "5.5 - 17# TSH W PIN"
    assert sheet["B13"].value == "5.510"
    assert sheet["B14"].value == "4.792"
    assert sheet["B15"].value == "120.125"
    assert sheet["B22"].value == "242K"
    assert sheet["B23"].value == "296K"
    assert sheet["B24"].value == "7,740"
    assert sheet["B25"].value == "6,290"
    assert sheet["B33"].value == "NA"
    assert sheet["B35"].value == "NA"
    assert sheet["B36"].value == "NA"
    assert sheet["B37"].value == "NA"
    assert sheet["H22"].value == "5.510 / 5.490"
    assert sheet["H23"].value == "4.812 / 4.792"
    assert sheet["H24"].value == "3.400 +.125/ -.000"
    assert sheet["H25"].value == 30
    assert sheet["H26"].value == "3.400 +.125/ -.000"
    assert sheet["H27"].value == 30
    workbook.close()
    tmp.cleanup()

    print("service tsh flow ok")


if __name__ == "__main__":
    main()
