"""Smoke check for VAM adapter integration in the generation service."""

from __future__ import annotations

import sys
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any

import fitz
from openpyxl import Workbook, load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from src.services.template_generation_service import (  # noqa: E402
    GenerationRequest,
    TemplateGenerationService,
)


class FakeVamAdapter:
    instances: list["FakeVamAdapter"] = []

    def __init__(
        self,
        base_url: str,
        configurator_url: str,
        logs_dir: str | Path,
        headless: bool,
        slow_mo: int,
        timeout_ms: int,
    ) -> None:
        self.base_url = base_url
        self.configurator_url = configurator_url
        self.logs_dir = Path(logs_dir)
        self.headless = headless
        self.slow_mo = slow_mo
        self.timeout_ms = timeout_ms
        self.closed = False
        self.run_calls: list[dict[str, Any]] = []
        self.instances.append(self)

    def run(self, mapped_result: dict[str, Any]) -> dict[str, Any]:
        self.run_calls.append(mapped_result)
        return {
            "tensile": "1000",
            "compression": "900",
            "burst": "800",
            "collapse": "700",
            "od": {"nominal": "5.500", "tol_1": "+0.010", "tol_2": "-0.010"},
            "id": {"nominal": "4.892", "tol_1": "+0.010", "tol_2": "-0.010"},
            "external_length": "7.250",
            "internal_length": "6.125",
            "drift": "4.767",
        }

    def close(self) -> None:
        self.closed = True


def main() -> None:
    FakeVamAdapter.instances.clear()
    text = (
        "POTS Document number: 123 Rev: A\n"
        "CP Part Number ABC-001\n"
        "Product Description Pup Joint 13CR(80) 5.5 17# VAM TOP BOX X "
        "5.5 17# TSH WEDGE PIN OAL 120\n"
        "ANSI/NACE MR0175/ISO 15156 (Yes/No) Yes\n"
        "QCP (Standard/Client Specific) Standard\n"
    )

    tmp = TemporaryDirectory()
    root = Path(tmp.name)
    pdf = root / "input.pdf"
    template = root / "template.xlsx"
    output_dir = root / "out"
    logs_dir = root / "logs"

    doc = fitz.open()
    page = doc.new_page()
    page.insert_text((72, 72), text)
    doc.save(pdf)
    doc.close()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Target"
    workbook.save(template)

    service = TemplateGenerationService(
        adapter_factories={"VAM": FakeVamAdapter},
        logs_dir=logs_dir,
    )
    request = GenerationRequest(
        input_path=pdf,
        template_path=template,
        output_dir=output_dir,
        target_sheet_name="Target",
        user_name="Tester",
        run_partner_adapters=True,
    )

    result = service.generate(request)
    output = Path(result.output_file)
    assert output.exists()

    assert len(FakeVamAdapter.instances) == 1
    adapter = FakeVamAdapter.instances[0]
    assert adapter.closed is True
    assert adapter.base_url == "https://www.vamservices.com"
    assert adapter.configurator_url == "https://www.vamservices.com/product/configurator"
    assert adapter.headless is True
    assert adapter.logs_dir == logs_dir

    assert len(adapter.run_calls) == 1
    mapped = adapter.run_calls[0]
    assert mapped["partner"] == "VAM"
    assert mapped["side"] == "upper"
    assert mapped["connection"]["name"] == "TOP"
    assert mapped["connection"]["od"] == "5-1/2"

    assert result.top_adapter == {
        "tensile": "1000",
        "compression": "900",
        "burst": "800",
        "collapse": "700",
        "od": {"nominal": "5.500", "tol_1": "+0.010", "tol_2": "-0.010"},
        "id": {"nominal": "4.892", "tol_1": "+0.010", "tol_2": "-0.010"},
        "external_length": "7.250",
        "internal_length": "6.125",
        "drift": "4.767",
    }
    assert result.bottom_adapter is None
    assert [item["partner"] for item in result.mapped_results or []] == ["VAM", "TSH"]

    formatted = result.writer_result["formatted"]
    assert formatted["top_thread"]["burst"] == "800"
    assert formatted["top_thread"]["drift"] == "4.767"
    assert "burst" not in formatted["bottom_thread"]

    workbook = load_workbook(output)
    sheet = workbook["Target"]
    assert sheet["B6"].value == "ABC-001"
    assert sheet["B28"].value == "5.5 - 17# VAM TOP BOX"
    assert sheet["B30"].value == "5.5 - 17# TSH W PIN"
    workbook.close()
    tmp.cleanup()

    print("service vam flow ok")


if __name__ == "__main__":
    main()
