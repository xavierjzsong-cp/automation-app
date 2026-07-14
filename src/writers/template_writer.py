"""Excel template writer for parsed POTS data."""

from __future__ import annotations

from dataclasses import asdict, is_dataclass
from datetime import date
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any
import logging
import re

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


logger = logging.getLogger(__name__)


class TemplateWriter:
    """Write parsed POTS fields into a selected Excel template sheet."""

    DEFAULT_ANGLE = 30
    NA = "NA"
    NA_VALUES = {"NA", "N/A"}

    def write(
        self,
        parsed: Any,
        top_adapter: dict[str, Any] | None = None,
        bottom_adapter: dict[str, Any] | None = None,
        template_path: str | Path | None = None,
        output_dir: str | Path | None = None,
        user_name: str | None = None,
        coating_data: dict[str, Any] | None = None,
        target_sheet_name: str | None = None,
    ) -> dict[str, Any]:
        """Write available parsed fields to the requested target sheet."""
        if template_path is None:
            raise ValueError("template_path is required.")
        if output_dir is None:
            raise ValueError("output_dir is required.")

        parsed_data = self._to_mapping(parsed)
        formatted = self._build_template_fields(
            parsed=parsed_data,
            top_adapter=top_adapter,
            bottom_adapter=bottom_adapter,
            user_name=user_name,
            coating_data=coating_data,
        )

        output_file = self._write_to_template(
            template_path=template_path,
            formatted=formatted,
            output_dir=output_dir,
            target_sheet_name=target_sheet_name,
        )

        return {
            "parsed": parsed_data,
            "formatted": formatted,
            "output_file": str(output_file),
            "target_sheet_name": target_sheet_name,
        }

    def _build_template_fields(
        self,
        parsed: dict[str, Any],
        top_adapter: dict[str, Any] | None,
        bottom_adapter: dict[str, Any] | None,
        user_name: str | None,
        coating_data: dict[str, Any] | None,
    ) -> dict[str, Any]:
        connections = parsed.get("connections") or {}
        upper = connections.get("upper") or {}
        lower = connections.get("lower") or {}

        top_thread_text = self._format_thread(
            od=self._get_value(upper, "od"),
            weight=self._get_value(upper, "weight"),
            family=self._get_value(upper, "family"),
            connection_name=self._get_value(upper, "name"),
            connection_type=self._get_value(upper, "type"),
        )
        bottom_thread_text = self._format_thread(
            od=self._get_value(lower, "od"),
            weight=self._get_value(lower, "weight"),
            family=self._get_value(lower, "family"),
            connection_name=self._get_value(lower, "name"),
            connection_type=self._get_value(lower, "type"),
        )

        material = self._format_material(
            mds=parsed.get("ansi_nace"),
            grade=parsed.get("product_material_grade"),
        )

        return {
            "user_name": self._format_user_name(user_name),
            "part_number": parsed.get("part_number"),
            "rev": parsed.get("rev"),
            "qcp": self._format_qcp(parsed.get("qcp")),
            "product_type": parsed.get("product_type"),
            "description": self._format_description(
                top_thread=top_thread_text,
                bottom_thread=bottom_thread_text,
                material=material,
            ),
            "material": material,
            "overall_length": self._format_overall_length(parsed.get("overall_length")),
            "coating": coating_data or {},
            "top_thread": {
                "thread": top_thread_text,
                **(top_adapter or {}),
            },
            "bottom_thread": {
                "thread": bottom_thread_text,
                **(bottom_adapter or {}),
            },
        }

    def _write_to_template(
        self,
        template_path: str | Path,
        formatted: dict[str, Any],
        output_dir: str | Path,
        target_sheet_name: str | None,
    ) -> Path:
        template_path = Path(template_path)
        output_dir = Path(output_dir)

        if not template_path.exists():
            raise FileNotFoundError(f"Template file not found: {template_path}")

        output_dir.mkdir(parents=True, exist_ok=True)
        workbook = load_workbook(template_path)
        sheet = self._get_target_sheet(workbook, target_sheet_name)

        self._write_if_editable(sheet, "G1", formatted.get("user_name"))
        self._write_if_editable(sheet, "B6", formatted.get("part_number"))
        self._write_if_editable(sheet, "D6", formatted.get("rev"))
        self._write_if_editable(sheet, "B8", formatted.get("product_type"))
        self._write_if_editable(sheet, "B18", formatted.get("material"))
        self._write_if_editable(sheet, "B28", (formatted.get("top_thread") or {}).get("thread"))
        self._write_if_editable(sheet, "B30", (formatted.get("bottom_thread") or {}).get("thread"))
        self._write_if_editable(sheet, "B34", formatted.get("qcp"))
        self._write_if_editable(sheet, "H9", formatted.get("overall_length"))

        coating = formatted.get("coating") or {}
        self._write_if_editable(sheet, "B29", coating.get("top_thread_coating"))
        self._write_if_editable(sheet, "B31", coating.get("bottom_thread_coating"))
        self._write_if_editable(sheet, "B32", coating.get("body_coating"))

        self._write_if_editable(sheet, "B13", self._get_overall_od_max(formatted))
        self._write_if_editable(sheet, "B14", self._get_overall_id_min(formatted))
        self._write_if_editable(
            sheet,
            "B15",
            self._get_max_overall_length(formatted.get("overall_length")),
        )
        self._write_if_editable(
            sheet,
            "B22",
            self._get_min_thread_metric(formatted, "tensile", suffix_k=True),
        )
        self._write_if_editable(
            sheet,
            "B23",
            self._get_min_thread_metric(formatted, "compression", suffix_k=True),
        )
        self._write_if_editable(
            sheet,
            "B24",
            self._get_min_thread_metric(formatted, "burst"),
        )
        self._write_if_editable(
            sheet,
            "B25",
            self._get_min_thread_metric(formatted, "collapse"),
        )
        self._write_if_editable(sheet, "B33", self._get_drift_size_for_template(formatted))

        top_thread = formatted.get("top_thread") or {}
        self._write_if_editable(sheet, "H13", self._format_thread_dimension(top_thread.get("od")))
        self._write_if_editable(sheet, "H14", self._format_thread_dimension(top_thread.get("id")))
        self._write_if_editable(sheet, "H15", self._format_thread_length(top_thread.get("external_length")))
        self._write_if_editable(sheet, "H17", self._format_thread_length(top_thread.get("internal_length")))

        bottom_thread = formatted.get("bottom_thread") or {}
        self._write_if_editable(sheet, "H22", self._format_thread_dimension(bottom_thread.get("od")))
        self._write_if_editable(sheet, "H23", self._format_thread_dimension(bottom_thread.get("id")))
        self._write_if_editable(sheet, "H24", self._format_thread_length(bottom_thread.get("external_length")))
        self._write_if_editable(sheet, "H26", self._format_thread_length(bottom_thread.get("internal_length")))

        self._write_if_editable(sheet, "B35", self.NA)
        self._write_if_editable(sheet, "B36", self.NA)
        self._write_if_editable(sheet, "B37", self.NA)
        self._write_if_editable(sheet, "H16", self.DEFAULT_ANGLE)
        self._write_if_editable(sheet, "H18", self.DEFAULT_ANGLE)
        self._write_if_editable(sheet, "H25", self.DEFAULT_ANGLE)
        self._write_if_editable(sheet, "H27", self.DEFAULT_ANGLE)

        part_number = formatted.get("part_number")
        if not part_number:
            raise ValueError("part_number is missing, cannot name output file.")

        output_file = output_dir / f"{part_number}.xlsx"
        workbook.save(output_file)
        logger.info("Saved template output: %s", output_file)
        return output_file

    def _get_target_sheet(self, workbook: Any, target_sheet_name: str | None) -> Any:
        target_sheet_name = str(target_sheet_name or "").strip()
        if not target_sheet_name:
            raise ValueError("Target sheet name is required.")

        if target_sheet_name not in workbook.sheetnames:
            available_sheets = ", ".join(workbook.sheetnames)
            raise ValueError(
                f"Target sheet not found in template: {target_sheet_name}. "
                f"Available sheets: {available_sheets}"
            )

        return workbook[target_sheet_name]

    def _write_if_editable(self, sheet: Any, cell_ref: str, value: Any) -> None:
        if value is None:
            return
        if not self._is_cell_editable(sheet, cell_ref):
            return
        sheet[cell_ref] = value

    def _is_cell_editable(self, sheet: Any, cell_ref: str) -> bool:
        cell = sheet[cell_ref]
        if isinstance(cell, MergedCell):
            return False

        current_value = cell.value
        if isinstance(current_value, str):
            normalized = current_value.strip().upper()
            if normalized in self.NA_VALUES:
                return False

        return True

    def _format_thread(
        self,
        od: str | None,
        weight: str | None,
        family: str | None,
        connection_name: str | None,
        connection_type: str | None,
    ) -> str | None:
        if not od or not weight or not connection_name or not connection_type:
            return None

        connection_label = self._format_connection_label(
            family=family,
            connection_name=connection_name,
            connection_type=connection_type,
        )
        if not connection_label:
            return None

        return f"{od} - {weight}# {connection_label}"

    def _format_connection_label(
        self,
        family: str | None,
        connection_name: str | None,
        connection_type: str | None,
    ) -> str | None:
        normalized_name = self._normalize_connection_name(connection_name)
        normalized_type = str(connection_type or "").strip().upper()
        normalized_family = str(family or "").strip().upper()

        if not normalized_name or not normalized_type:
            return None

        parts: list[str] = []
        if normalized_family and normalized_family != "HT":
            parts.append(normalized_family)
        parts.append(normalized_name)
        parts.append(normalized_type)
        return " ".join(parts)

    def _normalize_connection_name(self, connection_name: str | None) -> str | None:
        if not connection_name:
            return None

        text = str(connection_name).strip()
        text = re.sub(r"\bWEDGE\b", "W", text, flags=re.IGNORECASE)
        text = re.sub(r"\bW\s+(\d+)\b", r"W\1", text, flags=re.IGNORECASE)
        text = re.sub(r"\s+", " ", text).strip()
        return text or None

    def _format_user_name(self, user_name: str | None) -> str | None:
        if not user_name:
            return None

        name = re.sub(r"\s+", " ", str(user_name).strip())
        if not name:
            return None

        today_text = date.today().strftime("%d-%m-%Y")
        return f"{name.upper()} ({today_text})"

    def _format_material(self, mds: str | None, grade: str | None) -> str | None:
        if not mds and not grade:
            return None

        normalized_grade = self._normalize_grade(grade) if grade else None
        if mds and normalized_grade:
            return f"{mds} ({normalized_grade})"
        if mds:
            return mds
        return normalized_grade

    def _normalize_grade(self, grade: str) -> str:
        text = grade.strip()
        match = re.match(r"^([A-Za-z0-9]+)\(([\d.]+)\)$", text)
        if match:
            material_family = match.group(1)
            strength = match.group(2)
            return f"{material_family}-{strength}KSI"
        return text

    def _format_qcp(self, qcp: str | None) -> str | None:
        if not qcp:
            return None

        text = qcp.strip()
        text = re.sub(r"\bSTANDARD\b", "STD", text, flags=re.IGNORECASE)
        return re.sub(r"\s+", " ", text).strip() or None

    def _format_description(
        self,
        top_thread: str | None,
        bottom_thread: str | None,
        material: str | None,
    ) -> str | None:
        thread_part = None
        if top_thread and bottom_thread:
            thread_part = f"{top_thread} x {bottom_thread}"
        elif top_thread:
            thread_part = top_thread
        elif bottom_thread:
            thread_part = bottom_thread

        if thread_part and material:
            return f"{thread_part}, {material}"
        return thread_part or material

    def _format_overall_length(self, overall_length: str | None) -> str | None:
        if not overall_length:
            return None

        match = re.search(r"(\d+(?:\.\d+)?)", str(overall_length))
        if not match:
            return None

        value = float(match.group(1))
        return f"{value:.3f} +/-.125"

    def _get_max_overall_length(self, formatted_overall_length: str | None) -> str | None:
        if not formatted_overall_length:
            return None

        match = re.search(
            r"(\d+(?:\.\d+)?)\s*\+/-\s*([+-]?(?:\d+(?:\.\d+)?|\.\d+))",
            formatted_overall_length,
            flags=re.IGNORECASE,
        )
        if not match:
            return None

        nominal = float(match.group(1))
        tolerance = float(match.group(2))
        return f"{nominal + tolerance:.3f}"

    def _format_thread_dimension(self, dimension: dict[str, Any] | None) -> str | None:
        if not dimension:
            return None

        if "min" in dimension and "max" in dimension:
            return self._format_min_max_dimension(dimension)

        if "nominal" in dimension and "tol_1" in dimension and "tol_2" in dimension:
            return self._format_nominal_tolerance_dimension(dimension)

        return None

    def _format_min_max_dimension(self, dimension: dict[str, Any]) -> str | None:
        min_value = dimension.get("min")
        max_value = dimension.get("max")
        if not min_value or not max_value:
            return None
        return f"{max_value} / {min_value}"

    def _format_nominal_tolerance_dimension(self, dimension: dict[str, Any]) -> str | None:
        nominal = dimension.get("nominal")
        tol_1 = dimension.get("tol_1")
        tol_2 = dimension.get("tol_2")
        if not nominal or not tol_1 or not tol_2:
            return None

        tol_1 = str(tol_1).strip()
        tol_2 = str(tol_2).strip()

        if tol_1.startswith("+") and tol_2.startswith("-"):
            upper_tol = tol_1
            lower_tol = tol_2
        elif tol_1.startswith("-") and tol_2.startswith("+"):
            upper_tol = tol_2
            lower_tol = tol_1
        else:
            try:
                t1 = float(tol_1.replace("+", ""))
                t2 = float(tol_2.replace("+", ""))
                if t1 >= t2:
                    upper_tol = tol_1
                    lower_tol = tol_2
                else:
                    upper_tol = tol_2
                    lower_tol = tol_1
            except Exception:
                upper_tol = tol_1
                lower_tol = tol_2

        return f"{nominal} {self._compact_tol(upper_tol)} /{self._compact_tol(lower_tol)}"

    def _compact_tol(self, tol: str | None) -> str | None:
        if not tol:
            return None

        text = str(tol).strip()
        if text.startswith("+0."):
            return "+." + text[3:]
        if text.startswith("-0."):
            return "-." + text[3:]
        return text

    def _format_thread_length(self, length_value: Any) -> str | None:
        value = self._extract_decimal_from_value(length_value)
        if value is None:
            return None
        return f"{self._format_decimal(value)} +.125/ -.000"

    def _format_drift_size(self, drift_value: Decimal) -> str:
        return f"{self._format_decimal(drift_value)} +.020/-.000"

    def _get_thread_dimension_max(self, dimension: dict[str, Any] | None) -> Decimal | None:
        if not dimension:
            return None

        if "min" in dimension and "max" in dimension:
            return self._to_decimal(dimension.get("max"))

        if "nominal" in dimension and "tol_1" in dimension and "tol_2" in dimension:
            nominal = self._to_decimal(dimension.get("nominal"))
            tol_1 = self._to_decimal(dimension.get("tol_1"))
            tol_2 = self._to_decimal(dimension.get("tol_2"))
            if nominal is None or tol_1 is None or tol_2 is None:
                return None
            return nominal + max(tol_1, tol_2)

        return None

    def _get_thread_dimension_min(self, dimension: dict[str, Any] | None) -> Decimal | None:
        if not dimension:
            return None

        if "min" in dimension and "max" in dimension:
            return self._to_decimal(dimension.get("min"))

        if "nominal" in dimension and "tol_1" in dimension and "tol_2" in dimension:
            nominal = self._to_decimal(dimension.get("nominal"))
            tol_1 = self._to_decimal(dimension.get("tol_1"))
            tol_2 = self._to_decimal(dimension.get("tol_2"))
            if nominal is None or tol_1 is None or tol_2 is None:
                return None
            return nominal + min(tol_1, tol_2)

        return None

    def _get_overall_od_max_decimal(self, formatted: dict[str, Any]) -> Decimal | None:
        top_thread = formatted.get("top_thread") or {}
        bottom_thread = formatted.get("bottom_thread") or {}
        candidates = [
            self._get_thread_dimension_max(top_thread.get("od")),
            self._get_thread_dimension_max(bottom_thread.get("od")),
        ]
        candidates = [value for value in candidates if value is not None]
        if not candidates:
            return None
        return max(candidates)

    def _get_overall_id_min_decimal(self, formatted: dict[str, Any]) -> Decimal | None:
        top_thread = formatted.get("top_thread") or {}
        bottom_thread = formatted.get("bottom_thread") or {}
        candidates = [
            self._get_thread_dimension_min(top_thread.get("id")),
            self._get_thread_dimension_min(bottom_thread.get("id")),
        ]
        candidates = [value for value in candidates if value is not None]
        if not candidates:
            return None
        return min(candidates)

    def _get_overall_od_max(self, formatted: dict[str, Any]) -> str | None:
        value = self._get_overall_od_max_decimal(formatted)
        if value is None:
            return None
        return self._format_decimal(value)

    def _get_overall_id_min(self, formatted: dict[str, Any]) -> str | None:
        value = self._get_overall_id_min_decimal(formatted)
        if value is None:
            return None
        return self._format_decimal(value)

    def _get_min_thread_metric(
        self,
        formatted: dict[str, Any],
        metric_name: str,
        suffix_k: bool = False,
    ) -> str | None:
        top_thread = formatted.get("top_thread") or {}
        bottom_thread = formatted.get("bottom_thread") or {}
        candidates: list[Decimal] = []

        top_value = self._extract_decimal_from_value(top_thread.get(metric_name))
        bottom_value = self._extract_decimal_from_value(bottom_thread.get(metric_name))
        if top_value is not None:
            candidates.append(top_value)
        if bottom_value is not None:
            candidates.append(bottom_value)
        if not candidates:
            return None

        min_value = min(candidates)
        if suffix_k:
            return f"{self._format_metric_number(min_value)}K"
        return self._format_metric_number(min_value, use_comma=True)

    def _get_drift_size_for_template(self, formatted: dict[str, Any]) -> str:
        top_thread = formatted.get("top_thread") or {}
        bottom_thread = formatted.get("bottom_thread") or {}

        top_drift_raw = top_thread.get("drift")
        if self._is_na_value(top_drift_raw):
            return self.NA

        id_min = self._get_overall_id_min_decimal(formatted)
        if id_min is None:
            return self.NA

        top_drift = self._extract_decimal_from_value(top_drift_raw)
        if top_drift is not None and top_drift < id_min:
            return self._format_drift_size(top_drift)

        bottom_drift_raw = bottom_thread.get("drift")
        if self._is_na_value(bottom_drift_raw):
            return self.NA

        bottom_drift = self._extract_decimal_from_value(bottom_drift_raw)
        if bottom_drift is not None and bottom_drift < id_min:
            return self._format_drift_size(bottom_drift)

        return self.NA

    def _is_na_value(self, value: Any) -> bool:
        if value is None:
            return True
        return str(value).strip().upper() in {"", "NA", "N/A"}

    def _extract_decimal_from_value(self, value: Any) -> Decimal | None:
        if value is None:
            return None

        text = str(value).strip()
        if not text or self._is_na_value(text):
            return None

        match = re.search(r"[-+]?\d[\d,]*(?:\.\d+)?", text)
        if not match:
            return None
        return self._to_decimal(match.group(0))

    def _format_metric_number(self, value: Decimal, use_comma: bool = False) -> str:
        if value == value.to_integral_value():
            integer_value = int(value)
            if use_comma:
                return f"{integer_value:,}"
            return str(integer_value)

        normalized = value.normalize()
        text = format(normalized, "f")
        if not use_comma:
            return text

        integer_part, _, decimal_part = text.partition(".")
        integer_part = f"{int(integer_part):,}"
        return f"{integer_part}.{decimal_part}" if decimal_part else integer_part

    def _to_decimal(self, value: Any) -> Decimal | None:
        if value is None:
            return None

        text = str(value).strip().replace(",", "")
        if not text:
            return None

        try:
            return Decimal(text)
        except InvalidOperation:
            return None

    def _format_decimal(self, value: Decimal) -> str:
        value = value.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)
        return f"{value:.3f}"

    def _to_mapping(self, value: Any) -> dict[str, Any]:
        if isinstance(value, dict):
            return value
        if hasattr(value, "to_dict"):
            return value.to_dict()
        if is_dataclass(value):
            return asdict(value)
        raise TypeError(f"Unsupported parsed data type: {type(value)!r}")

    def _get_value(self, source: Any, key: str) -> Any:
        if isinstance(source, dict):
            return source.get(key)
        return getattr(source, key, None)
